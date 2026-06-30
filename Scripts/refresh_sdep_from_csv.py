#!/usr/bin/env python3
"""
Refresh SDEP internal DB_* sheets from external CSV files.

Usage from repo root:
    python Scripts/refresh_sdep_from_csv.py

This produces:
    Excel/Current/SDEP_v4_2_csv_import_refreshed.xlsx

Requires:
    pip install openpyxl
"""

from pathlib import Path
import csv
import shutil
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Excel" / "Current" / "SDEP_v4_2_csv_import.xlsx"
OUTPUT = ROOT / "Excel" / "Current" / "SDEP_v4_2_csv_import_refreshed.xlsx"

CSV_TO_SHEET = {
    ROOT / "Data" / "Panels.csv": "DB_Panels",
    ROOT / "Data" / "Inverters.csv": "DB_Inverters",
    ROOT / "Data" / "Batteries.csv": "DB_Batteries",
    ROOT / "Data" / "Lists.csv": "DB_Lists",
}

def parse_value(value: str):
    if value is None:
        return None
    s = value.strip()
    if s == "":
        return None
    # Preserve text with leading zeros or obvious non-numeric content.
    try:
        if any(ch in s for ch in [".", "e", "E"]):
            return float(s)
        return int(s)
    except ValueError:
        return s

def clear_sheet(ws):
    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None

def import_csv(ws, csv_path):
    clear_sheet(ws)
    with csv_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).value = parse_value(value)

def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")
    wb = load_workbook(WORKBOOK)
    for csv_path, sheet_name in CSV_TO_SHEET.items():
        if not csv_path.exists():
            raise FileNotFoundError(f"CSV not found: {csv_path}")
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        import_csv(wb[sheet_name], csv_path)
        print(f"Imported {csv_path.name} -> {sheet_name}")
    wb.save(OUTPUT)
    print(f"Saved refreshed workbook: {OUTPUT}")

if __name__ == "__main__":
    main()
